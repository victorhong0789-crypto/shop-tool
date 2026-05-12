import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

# -------------------------- 页面基础配置 --------------------------
st.set_page_config(
    page_title="店铺周度运营&广告数据分析系统",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# -------------------------- 样式常量定义（Excel美化） --------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal="right", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
PREMIUM_ROW_FILL = PatternFill("solid", fgColor="E6F4EA")
LOW_EFFICIENCY_ROW_FILL = PatternFill("solid", fgColor="FFF9E6")
SERIOUS_LOW_ROW_FILL = PatternFill("solid", fgColor="FCE4E4")
ALTERNATE_ROW_FILL = PatternFill("solid", fgColor="F2F2F2")

# -------------------------- 侧边栏：参数配置 --------------------------
with st.sidebar:
    st.header("⚙️ 行业参考值配置")
    st.caption("可灵活调整各项指标基准，系统自动按新规则计算")
    
    st.subheader("一、核心指标合格基准")
    shop_ctr_benchmark = st.slider("全商品平均点击率合格基准(≥%)", 1.0, 10.0, 3.0, 0.1)
    shop_cvr_benchmark = st.slider("全商品平均订单转化率合格基准(≥%)", 1.0, 15.0, 6.0, 0.1)
    ad_roas_benchmark = st.slider("广告商品整体ROAS合格基准(≥)", 1.0, 30.0, 15.0, 0.5)
    ad_cost_ratio_benchmark = st.slider("总广告花费占总销售额比例上限(≤%)", 1.0, 10.0, 5.0, 0.1)
    
    st.subheader("二、广告商品分级ROAS阈值")
    premium_roas = st.slider("优质广告商品ROAS下限(≥)", 10.0, 30.0, 20.0, 0.5)
    standard_roas_low = st.slider("达标广告商品ROAS下限(≥)", 5.0, 20.0, 10.0, 0.5)
    standard_roas_high = premium_roas
    low_efficiency_roas_low = st.slider("低效广告商品ROAS下限(≥)", 3.0, 15.0, 8.0, 0.5)
    low_efficiency_roas_high = standard_roas_low
    serious_low_roas = low_efficiency_roas_low
    
    st.subheader("三、潜力商品分级销售额阈值(PHP)")
    s_level_sales = st.slider("S级潜力商品周销售额下限(≥)", 1000, 30000, 10000, 500)
    a_level_sales_low = st.slider("A级潜力商品周销售额下限(≥)", 500, 20000, 5000, 500)
    a_level_sales_high = s_level_sales
    b_level_sales_low = st.slider("B级潜力商品周销售额下限(≥)", 100, 10000, 1000, 100)
    b_level_sales_high = a_level_sales_low
    
    st.header("📁 数据文件上传")
    st.caption("请上传以下2个Excel文件，完全匹配您的表格表头")
    shop_file = st.file_uploader("1. 店铺全商品表现表", type=["xlsx"], key="shop_file")
    ad_file = st.file_uploader("2. 广告商品表现表", type=["xlsx", "csv"], key="ad_file")

# -------------------------- 核心函数：数据加载与清洗 --------------------------
def load_and_validate_data(shop_file, ad_file):
    # 1. 读取店铺全商品表（Excel）
    shop_df = None
    try:
        shop_excel = pd.ExcelFile(shop_file)
        # 优先读取“热销商品”sheet，没有则读第一个sheet
        if "热销商品" in shop_excel.sheet_names:
            shop_df = pd.read_excel(shop_file, sheet_name="热销商品")
        else:
            shop_df = pd.read_excel(shop_file)
        
        # 检查并兼容“商品编号”列名（防止表头有空格或其他符号）
        shop_df.columns = shop_df.columns.str.strip()
        # 如果列名不是“商品编号”，自动找包含“商品编号”或“item_id”的列
        if "商品编号" not in shop_df.columns:
            possible_cols = [col for col in shop_df.columns if "商品编号" in col or "item_id" in str(col).lower()]
            if possible_cols:
                shop_df.rename(columns={possible_cols[0]: "商品编号"}, inplace=True)
            else:
                st.error("❌ 店铺全商品表中未找到“商品编号”列，请检查表头！")
                return None, None

        # 数据清洗函数
        def clean_numeric(s):
            if pd.isna(s) or s == "-":
                return 0
            s = str(s).replace(",", "").replace("%", "")
            try:
                return float(s)
            except:
                return 0
        
        # 清洗流量类数据
        traffic_cols = ["商品显示次数", "商品点击数", "不重复的商品展示次数", "不重复的商品点击数", 
                       "商品访客（访问）", "商品页面访客", "跳出商品页面的访客数", "搜索点击数", "赞"]
        for col in traffic_cols:
            if col in shop_df.columns:
                shop_df[col] = shop_df[col].apply(clean_numeric)
        
        # 清洗百分比数据
        percent_cols = ["点击率", "订单转化率（已下订单）", "订单转化率（已确认订单）", "转化率（已下单）", 
                       "转化率（已确认订单）", "商品跳出率", "转化率 (加入购物车率)", "重复下单率（已下订单）", 
                       "重复下单率（已确认订单）"]
        for col in percent_cols:
            if col in shop_df.columns:
                shop_df[col] = shop_df[col].apply(clean_numeric)
        
        # 清洗销售额数据
        sales_cols = ["销售额（已下单） (PHP)", "销售额（已确认订单） (PHP)", "每笔订单销售额（已下订单） (PHP)", 
                     "每笔订单销售额（已确认订单） (PHP)"]
        for col in sales_cols:
            if col in shop_df.columns:
                shop_df[col] = shop_df[col].apply(clean_numeric)
        
        # 清洗订单数据
        order_cols = ["已下订单", "已确定订单", "件数（已下单）", "件数（已确认订单）", "买家数（已下单）", 
                     "买家数（已确认订单）", "商品访客（添加至购物车）", "件数 (加入购物车）"]
        for col in order_cols:
            if col in shop_df.columns:
                shop_df[col] = shop_df[col].apply(clean_numeric)
        
        # 清洗重复下单数据
        repeat_cols = ["重复下单平均天数（已下订单）", "重复下单的平均天数（已确认订单）"]
        for col in repeat_cols:
            if col in shop_df.columns:
                shop_df[col] = shop_df[col].apply(clean_numeric)
        
        # 去重与空值处理
        shop_df = shop_df.drop_duplicates(subset=["商品编号"], keep="first")
        shop_df = shop_df.dropna(subset=["商品编号", "商品"]).reset_index(drop=True)
        
    except Exception as e:
        st.error(f"❌ 店铺全商品表读取失败: {str(e)}")
        return None, None
    
    # 2. 读取广告商品表（CSV）
    ad_df = pd.DataFrame()
    if ad_file is not None:
        try:
            # 万能 CSV 读取方案，解决格式混乱问题
            ad_df = pd.read_csv(
                ad_file,
                on_bad_lines='skip',       # 跳过格式错误的行
                engine='python',           # 使用更宽松的 Python 引擎
                skiprows=7,                # 跳过前7行无效标题
                encoding='utf-8',
                sep=None                   # 自动识别分隔符
            )
        
            # 清洗列名
            ad_df.columns = ad_df.columns.str.strip()
        
            # 检查商品编号列
            if "商品编号" not in ad_df.columns:
                possible_cols = [col for col in ad_df.columns if "商品编号" in col or "item_id" in str(col).lower()]
                if possible_cols:
                    ad_df.rename(columns={possible_cols[0]: "商品编号"}, inplace=True)
                else:
                    st.error("❌ 广告商品表中未找到“商品编号”列，请检查表头！")
                    return None, None
        
            # 清洗数值数据
            def clean_numeric(s):
                if pd.isna(s) or s == "-":
                    return 0
                s = str(s).replace(",", "").replace("%", "")
                try:
                    return float(s)
                except:
                    return 0
        
            ad_numeric_cols = ["展示次数", "点击数", "转化", "销售金额", "花费", "广告支出回报率"]
            for col in ad_numeric_cols:
                if col in ad_df.columns:
                    ad_df[col] = ad_df[col].apply(clean_numeric)
        
            # 清洗百分比数据
            ad_percent_cols = ["点击率", "转化率"]
            for col in ad_percent_cols:
                if col in ad_df.columns:
                    ad_df[col] = ad_df[col].apply(clean_numeric)
        
            # 去重与空值处理
            ad_df = ad_df.drop_duplicates(subset=["商品编号"], keep="first")
            if "广告名称" in ad_df.columns:
                ad_df = ad_df.dropna(subset=["商品编号", "广告名称"]).reset_index(drop=True)
            else:
                ad_df = ad_df.dropna(subset=["商品编号"]).reset_index(drop=True)
        
        except Exception as e:
            st.error(f"❌ 广告商品表读取失败: {str(e)}")
            return None, None
    
    return shop_df, ad_df
# -------------------------- 1. 核心数据汇总表 --------------------------
def generate_core_summary(shop_df, ad_df):
    total_skus = shop_df["商品编号"].nunique()
    total_impression = shop_df["商品显示次数"].sum()
    total_click = shop_df["商品点击数"].sum()
    total_order = shop_df["已确定订单"].sum()
    total_sales = shop_df["销售额（已确认订单） (PHP)"].sum()
    
    avg_ctr = (total_click / total_impression) * 100 if total_impression > 0 else 0
    avg_cvr = (total_order / total_click) * 100 if total_click > 0 else 0
    
    avg_asp = shop_df["每笔订单销售额（已确认订单） (PHP)"].mean() if "每笔订单销售额（已确认订单） (PHP)" in shop_df.columns else 0
    total_cart = shop_df["件数 (加入购物车）"].sum() if "件数 (加入购物车）" in shop_df.columns else 0
    cart_rate = shop_df["转化率 (加入购物车率)"].mean() if "转化率 (加入购物车率)" in shop_df.columns else 0
    repeat_rate = shop_df["重复下单率（已确认订单）"].mean() if "重复下单率（已确认订单）" in shop_df.columns else 0
    
    ad_total_cost = 0
    ad_total_impression = 0
    ad_total_click = 0
    ad_total_order = 0
    ad_total_sales = 0
    ad_overall_roas = 0
    ad_avg_ctr = 0
    ad_avg_cvr = 0
    ad_skus = 0
    hot_skus = shop_df[shop_df["已确定订单"] > 0]["商品编号"].nunique()
    ad_cover_ratio = 0
    ad_cost_ratio = 0
    
    if len(ad_df) > 0:
        ad_total_cost = ad_df["花费"].sum() if "花费" in ad_df.columns else 0
        ad_total_impression = ad_df["展示次数"].sum() if "展示次数" in ad_df.columns else 0
        ad_total_click = ad_df["点击数"].sum() if "点击数" in ad_df.columns else 0
        ad_total_order = ad_df["转化"].sum() if "转化" in ad_df.columns else 0
        ad_total_sales = ad_df["销售金额"].sum() if "销售金额" in ad_df.columns else 0
        ad_overall_roas = ad_total_sales / ad_total_cost if ad_total_cost > 0 else 0
        
        # ✅ 适配广告CSV：正确计算点击率、转化率
        ad_avg_ctr = (ad_total_click / ad_total_impression) * 100 if ad_total_impression > 0 else 0
        ad_avg_cvr = (ad_total_order / ad_total_click) * 100 if ad_total_click > 0 else 0
        
        ad_skus = ad_df["商品编号"].nunique()
        ad_cover_ratio = (ad_skus / hot_skus) * 100 if hot_skus > 0 else 0
        ad_cost_ratio = (ad_total_cost / total_sales) * 100 if total_sales > 0 else 0
    
    summary_data = []
    summary_data.append(["【全店整体运营表现】", "", "", ""])
    summary_data.append(["全店有效商品总数", f"{total_skus:,}", "-", "-"])
    summary_data.append(["全店总曝光量", f"{total_impression:,}", "-", "-"])
    summary_data.append(["全店总点击量", f"{total_click:,}", "-", "-"])
    summary_data.append(["全店总加购件数", f"{total_cart:,}", "-", "-"])
    summary_data.append(["全店总订单数", f"{total_order:,}", "-", "-"])
    summary_data.append(["全店总销售金额(PHP)", f"{total_sales:,.2f}", "-", "-"])
    summary_data.append(["全店平均点击率", f"{avg_ctr:.2f}%", f"≥{shop_ctr_benchmark:.1f}%", "达标" if avg_ctr >= shop_ctr_benchmark else "未达标"])
    summary_data.append(["全店平均订单转化率", f"{avg_cvr:.2f}%", f"≥{shop_cvr_benchmark:.1f}%", "达标" if avg_cvr >= shop_cvr_benchmark else "未达标"])
    summary_data.append(["全店平均加购率", f"{cart_rate:.2f}%", "≥5%", "达标" if cart_rate >= 5 else "未达标"])
    summary_data.append(["全店平均复购率", f"{repeat_rate:.2f}%", "≥3%", "达标" if repeat_rate >= 3 else "未达标"])
    summary_data.append(["全店平均客单价(PHP)", f"{avg_asp:,.2f}", "-", "-"])
    summary_data.append(["", "", "", ""])
    
    summary_data.append(["【广告投放整体表现】", "", "", ""])
    summary_data.append(["总广告花费(PHP)", f"{ad_total_cost:,.2f}", "-", "-"])
    summary_data.append(["总广告曝光量", f"{ad_total_impression:,}", "-", "-"])
    summary_data.append(["总广告点击量", f"{ad_total_click:,}", "-", "-"])
    summary_data.append(["总广告订单数", f"{ad_total_order:,}", "-", "-"])
    summary_data.append(["总广告销售金额(PHP)", f"{ad_total_sales:,.2f}", "-", "-"])
    summary_data.append(["整体广告ROAS", f"{ad_overall_roas:.2f}", f"≥{ad_roas_benchmark:.1f}", "达标" if ad_overall_roas >= ad_roas_benchmark else "未达标"])
    summary_data.append(["广告商品平均点击率", f"{ad_avg_ctr:.2f}%", "≥3.5%", "达标" if ad_avg_ctr >= 3.5 else "未达标"])
    summary_data.append(["广告商品平均转化率", f"{ad_avg_cvr:.2f}%", "≥8.0%", "达标" if ad_avg_cvr >= 8.0 else "未达标"])
    summary_data.append(["广告商品覆盖全店热销商品占比", f"{ad_cover_ratio:.2f}%", "≥30%", "达标" if ad_cover_ratio >= 30 else "未达标"])
    summary_data.append(["总广告花费占全店总销售额比例", f"{ad_cost_ratio:.2f}%", f"≤{ad_cost_ratio_benchmark:.1f}%", "达标" if ad_cost_ratio <= ad_cost_ratio_benchmark else "未达标"])
    
    summary_df = pd.DataFrame(summary_data, columns=["指标名称", "指标数值", "行业参考值", "达标状态"])
    return summary_df, {
        "total_sales": total_sales,
        "avg_ctr": avg_ctr,
        "avg_cvr": avg_cvr,
        "ad_total_cost": ad_total_cost,
        "ad_total_sales": ad_total_sales,
        "ad_overall_roas": ad_overall_roas,
        "ad_cover_ratio": ad_cover_ratio,
        "ad_cost_ratio": ad_cost_ratio
    }

# -------------------------- 2. 广告商品分析表 --------------------------
def generate_ad_analysis(ad_df):
    if len(ad_df) == 0:
        return pd.DataFrame(columns=["商品排名", "广告名称", "商品编号", "广告花费", "曝光量", "点击量", "转化数", "销售金额", "ROAS", "点击率", "转化率", "商品分级", "单商品优化建议"])
    
    ad_sorted_df = ad_df.sort_values(by="广告支出回报率", ascending=False).reset_index(drop=True)
    ad_sorted_df["商品排名"] = ad_sorted_df.index + 1
    
    def get_ad_level(row):
        roas = row["广告支出回报率"]
        cost = row["花费"]
        order = row["转化"]
        if order == 0:
            return "无效广告商品"
        elif roas >= premium_roas:
            return "优质广告商品"
        elif standard_roas_low <= roas < standard_roas_high:
            return "达标广告商品"
        elif low_efficiency_roas_low <= roas < low_efficiency_roas_high:
            return "低效广告商品"
        elif roas < serious_low_roas and cost > 100:
            return "严重低效广告商品"
        else:
            return "达标广告商品"
    
    ad_sorted_df["商品分级"] = ad_sorted_df.apply(get_ad_level, axis=1)
    
    # ✅ 适配CSV：点击率、转化率 正确计算 + 百分比格式
    ad_sorted_df["点击率"] = (ad_sorted_df["点击数"] / ad_sorted_df["展示次数"] * 100).round(2).astype(str) + "%"
    ad_sorted_df["转化率"] = (ad_sorted_df["转化"] / ad_sorted_df["点击数"] * 100).round(2).astype(str) + "%"
    
    def get_ad_suggestion(row):
        level = row["商品分级"]
        roas = row["广告支出回报率"]
        if level == "优质广告商品":
            return f"ROAS达{roas:.2f}，表现优异，建议加大预算投放，提升曝光量，打造核心爆款，同步优化素材进一步放大流量"
        elif level == "达标广告商品":
            return f"ROAS达{roas:.2f}，符合行业标准，建议维持当前预算，优化素材和详情页，进一步提升转化率和ROAS"
        elif level == "低效广告商品":
            return f"ROAS仅{roas:.2f}，表现偏低，建议缩减预算，优化商品详情页和定价，测试新的广告素材，提升转化能力"
        elif level == "严重低效广告商品":
            return f"ROAS仅{roas:.2f}，且广告花费超100PHP，严重低效，建议立即暂停投放，全面优化商品详情、定价和素材后，再小预算测试"
        elif level == "无效广告商品":
            return f"无任何转化订单，完全无效，建议立即暂停投放，排查商品定价、详情页、市场需求等核心问题，优化后再考虑投放"
        else:
            return "维持当前投放，持续监控数据变化"
    
    ad_sorted_df["单商品优化建议"] = ad_sorted_df.apply(get_ad_suggestion, axis=1)
    
    ad_analysis_df = ad_sorted_df[[
        "商品排名", "广告名称", "商品编号", "花费", "展示次数", "点击数",
        "转化", "销售金额", "广告支出回报率", "点击率", "转化率",
        "商品分级", "单商品优化建议"
    ]].rename(columns={
        "花费": "广告花费",
        "展示次数": "曝光量",
        "点击数": "点击量",
        "转化": "转化数",
        "广告支出回报率": "ROAS"
    })
    
    return ad_analysis_df

# -------------------------- 3. 潜力商品分析表 --------------------------
def generate_potential_analysis(shop_df, ad_df):
    ad_sku_list = ad_df["商品编号"].unique() if len(ad_df) > 0 else []
    potential_df = shop_df[~shop_df["商品编号"].isin(ad_sku_list)].copy()
    potential_sorted_df = potential_df.sort_values(by="销售额（已确认订单） (PHP)", ascending=False).reset_index(drop=True)
    potential_sorted_df["潜力排名"] = potential_sorted_df.index + 1
    
    # ✅ 修复：点击率 正确计算 + 百分比格式
    potential_sorted_df["点击率"] = (potential_sorted_df["商品点击数"] / potential_sorted_df["商品显示次数"] * 100).round(2).astype(str) + "%"
    potential_sorted_df["订单转化率（已确认订单）"] = (potential_sorted_df["已确定订单"] / potential_sorted_df["商品点击数"] * 100).round(2).astype(str) + "%"
    
    def get_potential_level(row):
        sales = row["销售额（已确认订单） (PHP)"]
        cvr = row["订单转化率（已确认订单）"]
        ctr = row["点击率"]
        if sales >= s_level_sales:
            return "S级潜力商品"
        elif a_level_sales_low <= sales < a_level_sales_high:
            return "A级潜力商品"
        elif b_level_sales_low <= sales < b_level_sales_high:
            return "B级潜力商品"
        else:
            return "普通商品"
    
    potential_sorted_df["商品分级"] = potential_sorted_df.apply(get_potential_level, axis=1)
    potential_filtered_df = potential_sorted_df[potential_sorted_df["商品分级"].isin(["S级潜力商品", "A级潜力商品", "B级潜力商品"])].copy()
    potential_top30_df = potential_filtered_df.head(30).reset_index(drop=True)
    potential_top30_df["潜力排名"] = potential_top30_df.index + 1
    
    def get_launch_priority(row):
        level = row["商品分级"]
        if level == "S级潜力商品":
            return "最高优先级"
        elif level == "A级潜力商品":
            return "高优先级"
        elif level == "B级潜力商品":
            return "中优先级"
        else:
            return "低优先级"
    
    potential_top30_df["投放优先级"] = potential_top30_df.apply(get_launch_priority, axis=1)
    
    def get_launch_suggestion(row):
        level = row["商品分级"]
        sales = row["销售额（已确认订单） (PHP)"]
        if level == "S级潜力商品":
            return f"周销售额达{sales:,.2f}PHP，转化和点击率均表现优异，自然转化能力极强，建议优先制作高质量广告素材，直接加大预算投放，快速放大销量，打造核心爆款"
        elif level == "A级潜力商品":
            return f"周销售额达{sales:,.2f}PHP，转化表现良好，有较大增长潜力，建议采用小预算测试投放，优化素材和出价，观察转化效果，若ROAS达标可逐步加大预算"
        elif level == "B级潜力商品":
            return f"周销售额达{sales:,.2f}PHP，有一定的市场需求，建议先优化商品详情页、定价和主图，提升自然转化能力后，再考虑小预算投放广告测试效果"
        else:
            return "先优化商品基础信息，提升自然转化后，再考虑投放"
    
    potential_top30_df["投放建议"] = potential_top30_df.apply(get_launch_suggestion, axis=1)
    
    potential_analysis_df = potential_top30_df[[
        "潜力排名", "商品", "商品编号", "商品显示次数", "商品点击数", "已确定订单",
        "销售额（已确认订单） (PHP)", "点击率", "订单转化率（已确认订单）", "商品分级", "投放优先级", "投放建议"
    ]]
    
    return potential_analysis_df

# -------------------------- 4. 周度优化建议表 --------------------------
def generate_optimization_suggestion(core_metrics, ad_analysis_df, potential_analysis_df):
    core_conclusion = f"本周店铺总销售额PHP {core_metrics['total_sales']:,.2f}，广告投放ROAS达{core_metrics['ad_overall_roas']:.2f}，整体投放效果{'优异' if core_metrics['ad_overall_roas'] >= ad_roas_benchmark else '达标' if core_metrics['ad_overall_roas'] >= 6 else '不佳'}；全店平均点击率{core_metrics['avg_ctr']:.2f}%，{'达标' if core_metrics['avg_ctr'] >= shop_ctr_benchmark else '未达标'}；平均转化率{core_metrics['avg_cvr']:.2f}%，{'达标' if core_metrics['avg_cvr'] >= shop_cvr_benchmark else '未达标'}；广告商品仅覆盖{core_metrics['ad_cover_ratio']:.2f}%的热销商品，有较大的拓展空间。"
    
    suggestion_data = []
    suggestion_data.append(["一、核心数据结论", core_conclusion])
    suggestion_data.append(["", ""])
    suggestion_data.append(["二、分模块优化建议", ""])
    suggestion_data.append(["（一）广告投放结构优化", ""])
    
    premium_skus = ad_analysis_df[ad_analysis_df["商品分级"] == "优质广告商品"]["广告名称"].tolist() if len(ad_analysis_df) > 0 else []
    premium_sku_str = "、".join(premium_skus[:5]) if len(premium_skus) > 0 else "无"
    suggestion_data.append(["1. 优质商品加大投放", f"针对ROAS TOP5的优质商品（{premium_sku_str}等，ROAS均超{premium_roas:.1f}），建议加大预算投放，提升曝光量，打造核心爆款；对ROAS {standard_roas_low:.1f}-{standard_roas_high:.1f}的表现良好商品，维持当前预算，优化素材进一步提升转化率。"])
    
    serious_low_skus = ad_analysis_df[ad_analysis_df["商品分级"] == "严重低效广告商品"]["广告名称"].tolist() if len(ad_analysis_df) > 0 else []
    serious_low_sku_str = "、".join(serious_low_skus[:5]) if len(serious_low_skus) > 0 else "无"
    invalid_skus = ad_analysis_df[ad_analysis_df["商品分级"] == "无效广告商品"]["广告名称"].count() if len(ad_analysis_df) > 0 else 0
    suggestion_data.append(["2. 低效商品及时止损", f"针对高花费低ROAS商品（{serious_low_sku_str}等，花费>100PHP，ROAS<{serious_low_roas:.1f}），建议立即缩减预算或暂停投放，重新优化商品详情页、定价和素材；对{invalid_skus}个无转化的商品，立即暂停投放，避免无效花费。"])
    
    suggestion_data.append(["3. 扩大广告覆盖范围", f"当前广告商品仅覆盖{core_metrics['ad_cover_ratio']:.2f}%的热销商品，远低于30%的行业参考值，建议优先从TOP30潜力商品中选择合适的商品拓展投放，挖掘更多增长机会。"])
    suggestion_data.append(["", ""])
    suggestion_data.append(["（二）潜力商品挖掘与投放", ""])
    
    s_level_skus = potential_analysis_df[potential_analysis_df["商品分级"] == "S级潜力商品"]["商品"].tolist()
    s_level_sku_str = "、".join(s_level_skus[:3]) if len(s_level_skus) > 0 else "无"
    suggestion_data.append(["1. S级潜力商品优先投放", f"对销售额≥{s_level_sales:,}PHP、转化率≥10%、点击率≥3%的S级潜力商品（{s_level_sku_str}等），这些商品自然转化表现优异，建议优先制作广告素材投放，快速放大销量。"])
    
    a_level_count = potential_analysis_df[potential_analysis_df["商品分级"] == "A级潜力商品"]["商品"].count()
    suggestion_data.append(["2. A级潜力商品测试投放", f"对销售额{a_level_sales_low:,}-{a_level_sales_high:,}PHP、转化率≥5%的{a_level_count}个A级潜力商品，采用小预算测试投放，优化素材和出价，观察转化效果，若ROAS达标可逐步加大预算。"])
    
    b_level_count = potential_analysis_df[potential_analysis_df["商品分级"] == "B级潜力商品"]["商品"].count()
    suggestion_data.append(["3. B级潜力商品先优化后投放", f"对销售额{b_level_sales_low:,}-{b_level_sales_high:,}PHP的{b_level_count}个B级潜力商品，可先优化商品详情和定价，提升自然转化后，再考虑小预算投放广告。"])
    suggestion_data.append(["", ""])
    suggestion_data.append(["（三）全店商品运营优化", ""])
    suggestion_data.append(["1. 优化广告商品落地页", "非广告商品的订单转化率显著高于广告商品，说明广告商品的落地页转化能力不足，建议针对广告商品优化详情页，突出产品卖点、用户评价、售后保障等内容，缩小与自然转化的差距。"])
    suggestion_data.append(["2. 通过广告引流带动全店转化", "广告商品的点击率优于非广告商品，引流精准度高，建议在广告商品的详情页中搭配相关商品推荐，通过广告引流带动全店其他商品的曝光和转化。"])
    suggestion_data.append(["3. 定期更新商品榜单", "每周更新热销商品榜单，对新进入热销榜的商品，及时评估是否适合投放广告，提前布局流量，抓住增长机会。"])
    suggestion_data.append(["", ""])
    suggestion_data.append(["（四）后续优化动作时间节点", ""])
    suggestion_data.append(["1. 3天内", "完成低效广告的暂停/预算缩减，完成优质商品的预算提升"])
    suggestion_data.append(["2. 7天内", "完成TOP10潜力商品的广告素材制作和投放测试"])
    suggestion_data.append(["3. 每周", "复盘广告数据，动态调整预算和出价，持续优化ROAS"])
    suggestion_data.append(["4. 每月", "更新潜力商品清单，扩大广告覆盖的商品范围，提升全店销售额"])
    suggestion_data.append(["", ""])
    
    risk_tips = []
    if core_metrics["ad_overall_roas"] < ad_roas_benchmark and len(ad_analysis_df) > 0:
        risk_tips.append(f"整体广告ROAS{core_metrics['ad_overall_roas']:.2f}，低于行业基准值{ad_roas_benchmark:.1f}，需重点管控广告花费，及时止损低效投放")
    if core_metrics["avg_ctr"] < shop_ctr_benchmark:
        risk_tips.append(f"全店平均点击率{core_metrics['avg_ctr']:.2f}%，低于行业基准值{shop_ctr_benchmark:.1f}%，需重点优化商品主图和标题，提升点击率")
    if core_metrics["avg_cvr"] < shop_cvr_benchmark:
        risk_tips.append(f"全店平均转化率{core_metrics['avg_cvr']:.2f}%，低于行业基准值{shop_cvr_benchmark:.1f}%，需重点优化商品详情页和定价，提升转化能力")
    if core_metrics["ad_cost_ratio"] > ad_cost_ratio_benchmark and len(ad_analysis_df) > 0:
        risk_tips.append(f"广告花费占销售额比例{core_metrics['ad_cost_ratio']:.2f}%，高于行业上限{ad_cost_ratio_benchmark:.1f}%，需管控广告投入，提升投放ROAS")
    if len(risk_tips) == 0:
        risk_tips.append("本周整体运营和投放表现良好，无重大风险，需持续监控数据变化，及时调整策略")
    suggestion_data.append(["三、风险提示", "\n".join(risk_tips)])
    
    suggestion_df = pd.DataFrame(suggestion_data, columns=["模块/动作", "详细内容"])
    return suggestion_df

# -------------------------- 5. 需要优化listing的商品表 --------------------------
def generate_listing_optimization(shop_df, ad_df):
    ad_sku_list = ad_df["商品编号"].unique() if len(ad_df) > 0 else []

    listing_df = shop_df.copy()
    listing_df["商品类型"] = listing_df["商品编号"].apply(
        lambda x: "广告商品" if x in ad_sku_list else "自然流商品"
    )

    # ✅ 修复：点击率、转化率 正确计算
    ctr = (listing_df["商品点击数"] / listing_df["商品显示次数"] * 100).round(2)
    cvr = (listing_df["已确定订单"] / listing_df["商品点击数"] * 100).round(2)
    listing_df["点击率"] = ctr.astype(str) + "%"
    listing_df["订单转化率（已确认订单）"] = cvr.astype(str) + "%"

    # 筛选低于基准值的商品
    listing_df = listing_df[
        (ctr < shop_ctr_benchmark) | 
        (cvr < shop_cvr_benchmark)
    ].copy()

    listing_df = listing_df[[
        "商品", "商品编号", "商品类型", "商品显示次数", "商品点击数",
        "已确定订单", "销售额（已确认订单） (PHP)", "点击率", "订单转化率（已确认订单）"
    ]].reset_index(drop=True)

    # 排序权重
    listing_df["问题严重度"] = listing_df.apply(
        lambda x: (1 if x["商品类型"] == "广告商品" else 0) * 1000
        + x["已确定订单"] * 10,
        axis=1
    )

    listing_sorted_df = listing_df.sort_values(by="问题严重度", ascending=False).reset_index(drop=True)
    listing_sorted_df["优化排名"] = listing_sorted_df.index + 1
    listing_top30_df = listing_sorted_df.head(30).copy()

    def get_core_issue(row):
        ctr_val = float(str(row["点击率"]).replace("%", ""))
        cvr_val = float(str(row["订单转化率（已确认订单）"]).replace("%", ""))
        issues = []
        if ctr_val < shop_ctr_benchmark:
            issues.append(f"点击率{ctr_val:.2f}% 低于 {shop_ctr_benchmark:.1f}%")
        if cvr_val < shop_cvr_benchmark:
            issues.append(f"转化率{cvr_val:.2f}% 低于 {shop_cvr_benchmark:.1f}%")
        return "｜".join(issues) if issues else "无问题"

    listing_top30_df["核心问题诊断"] = listing_top30_df.apply(get_core_issue, axis=1)

    def get_optimization_tip(row):
        ctr_val = float(str(row["点击率"]).replace("%", ""))
        cvr_val = float(str(row["订单转化率（已确认订单）"]).replace("%", ""))
        tips = []
        if ctr_val < shop_ctr_benchmark:
            tips.append("主图/标题优化")
        if cvr_val < shop_cvr_benchmark:
            tips.append("详情/定价/评价优化")
        return "｜".join(tips) if tips else "维持现状"

    listing_top30_df["针对性优化建议"] = listing_top30_df.apply(get_optimization_tip, axis=1)

    listing_optimization_df = listing_top30_df[[
        "优化排名", "商品", "商品编号", "商品类型", "商品显示次数", "商品点击数",
        "已确定订单", "销售额（已确认订单） (PHP)", "点击率", "订单转化率（已确认订单）",
        "核心问题诊断", "针对性优化建议"
    ]]

    return listing_optimization_df

# -------------------------- Excel生成 --------------------------
def generate_excel_file(summary_df, ad_analysis_df, potential_analysis_df, suggestion_df, listing_optimization_df):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    def create_sheet(wb, sheet_name, df):
        ws = wb.create_sheet(sheet_name)
        headers = df.columns.tolist()
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
        for row_idx, row_data in enumerate(df.values.tolist(), 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                if isinstance(value, (int, float)):
                    cell.alignment = RIGHT_ALIGN
                elif col_idx == 1:
                    cell.alignment = CENTER_ALIGN
                else:
                    cell.alignment = LEFT_ALIGN
                if row_idx % 2 == 0:
                    cell.fill = ALTERNATE_ROW_FILL
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    current_length = len(str(cell_value))
                    if current_length > max_length:
                        max_length = current_length
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        ws.freeze_panes = "A2"
        return ws
    
    create_sheet(wb, "核心数据汇总", summary_df)
    if len(ad_analysis_df) > 0:
        ad_ws = create_sheet(wb, "广告商品分析", ad_analysis_df)
        level_col_idx = ad_analysis_df.columns.get_loc("商品分级") + 1
        for row_idx in range(2, ad_ws.max_row + 1):
            level = ad_ws.cell(row=row_idx, column=level_col_idx).value
            fill = None
            if level == "优质广告商品":
                fill = PREMIUM_ROW_FILL
            elif level == "低效广告商品":
                fill = LOW_EFFICIENCY_ROW_FILL
            elif level == "严重低效广告商品" or level == "无效广告商品":
                fill = SERIOUS_LOW_ROW_FILL
            if fill:
                for col_idx in range(1, ad_ws.max_column + 1):
                    ad_ws.cell(row=row_idx, column=col_idx).fill = fill
    create_sheet(wb, "潜力商品分析", potential_analysis_df)
    suggestion_ws = create_sheet(wb, "周度优化建议", suggestion_df)
    suggestion_ws.column_dimensions["A"].width = 30
    suggestion_ws.column_dimensions["B"].width = 120
    create_sheet(wb, "需要优化listing的商品", listing_optimization_df)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# -------------------------- 主页面渲染 --------------------------
st.title("📊 跨境电商店铺周度运营&广告数据分析系统")
st.caption("100%匹配您的表格表头，标准化分析工具，支持灵活调整行业参考值，自动生成5张标准化分析表格")

if shop_file is not None:
    with st.spinner("正在加载并校验数据..."):
        shop_df, ad_df = load_and_validate_data(shop_file, ad_file)
    if shop_df is not None:
        st.success("✅ 数据加载成功，开始分析...")
        
        with st.spinner("正在生成核心数据汇总表..."):
            summary_df, core_metrics = generate_core_summary(shop_df, ad_df)
        with st.spinner("正在生成广告商品分析表..."):
            ad_analysis_df = generate_ad_analysis(ad_df)
        with st.spinner("正在生成潜力商品分析表..."):
            potential_analysis_df = generate_potential_analysis(shop_df, ad_df)
        with st.spinner("正在生成周度优化建议表..."):
            suggestion_df = generate_optimization_suggestion(core_metrics, ad_analysis_df, potential_analysis_df)
        with st.spinner("正在生成需要优化listing的商品表..."):
            listing_optimization_df = generate_listing_optimization(shop_df, ad_df)
        with st.spinner("正在生成Excel文件..."):
            excel_file = generate_excel_file(summary_df, ad_analysis_df, potential_analysis_df, suggestion_df, listing_optimization_df)
        
        st.success("✅ 分析完成，所有表格已生成！")
        
        st.header("📈 核心数据看板")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("全店总销售额(PHP)", f"{core_metrics['total_sales']:,.2f}")
        with col2:
            st.metric("整体广告ROAS", f"{core_metrics['ad_overall_roas']:.2f}", f"{'达标' if core_metrics['ad_overall_roas'] >= ad_roas_benchmark else '未达标'}")
        with col3:
            st.metric("全店平均点击率", f"{core_metrics['avg_ctr']:.2f}%", f"{'达标' if core_metrics['avg_ctr'] >= shop_ctr_benchmark else '未达标'}")
        with col4:
            st.metric("全店平均转化率", f"{core_metrics['avg_cvr']:.2f}%", f"{'达标' if core_metrics['avg_cvr'] >= shop_cvr_benchmark else '未达标'}")
        
        col5, col6, col7, col8 = st.columns(4)
        with col5:
            st.metric("总广告花费(PHP)", f"{core_metrics['ad_total_cost']:,.2f}")
        with col6:
            st.metric("广告销售占比", f"{core_metrics['ad_total_sales']/core_metrics['total_sales']*100:.2f}%" if core_metrics['total_sales'] > 0 else "0%")
        with col7:
            st.metric("广告花费占比", f"{core_metrics['ad_cost_ratio']:.2f}%", f"{'达标' if core_metrics['ad_cost_ratio'] <= ad_cost_ratio_benchmark else '未达标'}")
        with col8:
            st.metric("广告商品覆盖占比", f"{core_metrics['ad_cover_ratio']:.2f}%")
        
        st.header("📋 分析表格预览")
        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "核心数据汇总",
            "广告商品分析",
            "潜力商品分析",
            "周度优化建议",
            "需要优化listing的商品"
        ])
        with tab1:
            st.dataframe(summary_df, use_container_width=True, hide_index=True)
        with tab2:
            st.dataframe(ad_analysis_df, use_container_width=True, hide_index=True)
        with tab3:
            st.dataframe(potential_analysis_df, use_container_width=True, hide_index=True)
        with tab4:
            st.dataframe(suggestion_df, use_container_width=True, hide_index=True)
        with tab5:
            st.dataframe(listing_optimization_df, use_container_width=True, hide_index=True)
        
        st.header("📥 下载完整分析报告")
        st.download_button(
            label="下载Excel分析文件",
            data=excel_file,
            file_name="店铺周度运营&广告数据分析报告.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
else:
    st.info("👈 请在左侧边栏上传店铺全商品表现表，可选上传广告商品表现表，开始分析")
    st.subheader("📋 上传文件要求")
    st.markdown("""
    1. **店铺全商品表现表**：100%匹配您上传的表格表头，无需修改，直接上传即可
    2. **广告商品表现表**（可选）：匹配您上传的广告表表头，无需修改，直接上传即可
    3. 所有文件格式必须为.xlsx，无需修改字段名，直接上传即可
    """)