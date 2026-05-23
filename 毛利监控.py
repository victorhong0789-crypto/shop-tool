import streamlit as st
import pandas as pd
import math
from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ======================
# 苹果风全局样式（只加这段）
# ======================
st.markdown("""
<style>
    /* 全局背景 */
    .stApp {
        background-color: #f9fafb;
        font-family: -apple-system, BlinkMacSystemFont, 'San Francisco', 'Helvetica Neue', Helvetica, Arial, sans-serif;
    }
    /* 标题 */
    h1, h2, h3, h4, h5, h6 {
        color: #111827;
        font-weight: 600;
    }
    /* 卡片容器 */
    .block-container {
        padding-top: 2rem;
        padding-bottom: 2rem;
    }
    .st-card, div[data-testid="stMetric"], div[data-testid="stFileUploader"] {
        background: #ffffff;
        border-radius: 12px;
        padding: 1rem;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
        border: 1px solid #e5e7eb;
    }
    /* 按钮：苹果圆角浅蓝 */
    .stButton>button {
        background-color: #007aff;
        color: white;
        border-radius: 10px;
        border: none;
        padding: 0.5rem 1rem;
        font-weight: 500;
    }
    .stButton>button:hover {
        background-color: #0066cc;
    }
    /* 滑块 */
    .stSlider {
        padding: 0.5rem 0;
    }
    /* 表格 */
    .stDataFrame {
        border-radius: 12px;
        overflow: hidden;
        border: 1px solid #e5e7eb;
    }
    /* 提示信息 */
    .stInfo, .stSuccess, .stWarning, .stError {
        border-radius: 10px;
        border: none;
    }
</style>
""", unsafe_allow_html=True)

# ======================
# 页面配置
# ======================
st.set_page_config(page_title="毛利分析&自动提价工具", page_icon="📊", layout="wide")
st.title("📊 毛利监控")

# ======================
# 1. 毛利阈值 百分比显示
# ======================
profit_threshold_percent = st.slider(
    "筛选毛利阈值（低于该值将被筛选）",
    min_value=0, max_value=50, value=20, step=1, format="%d%%"
)
profit_threshold = profit_threshold_percent / 100
st.info(f"当前规则：只保留 毛利预估 < {profit_threshold_percent}% 的商品")

# ======================
# 2. 上传入口1：主数据文件
# ======================
st.subheader("📂 上传【商品运营毛利表】")
uploaded_file = st.file_uploader("上传主Excel文件", type=["xlsx"], key="main")

# ======================
# 3. 上传入口2：需要排除的店铺列表
# ======================
st.subheader("📂 上传【需要排除的店铺列表】")
exclude_shop_file = st.file_uploader("上传只含【店铺】一列的Excel", type=["xlsx"], key="exclude_shop")

# ======================
# 4. 上传入口3：需要排除的SKU列表
# ======================
st.subheader("📂 上传【需要排除的SKU列表】")
exclude_sku_file = st.file_uploader("上传只含【系统SKU】一列的Excel", type=["xlsx"], key="exclude_sku")

# ======================
# 5. 上传入口4：控价SKU列表
# ======================
st.subheader("📂 上传【控价SKU列表】")
price_control_sku_file = st.file_uploader("上传含【系统SKU】和【控价价格】两列的Excel", type=["xlsx"], key="price_control_sku")

# ======================
# 处理逻辑
# ======================
if uploaded_file is not None:
    df = pd.read_excel(uploaded_file)
    st.success("✅ 主文件上传成功！")

    # ----------------------
    # 自动识别列
    # ----------------------
    shop_col = next((c for c in df.columns if c in ["店铺", "店铺名称", "所属店铺", "店铺ID"]), None)
    sku_col = next((c for c in df.columns if c in ["系统SKU", "SKU", "商品SKU", "SKU编码"]), None)
    required_cols = ["销量", "商品优惠后金额（仅普通单）", "其他优惠", "商品采购成本"]

    # 检查缺失列
    missing = []
    if not shop_col: missing.append("店铺")
    if not sku_col: missing.append("系统SKU")
    for c in required_cols:
        if c not in df.columns: missing.append(c)
    if missing:
        st.error(f"❌ 缺少必要列：{', '.join(missing)}")
        st.stop()

    # ======================
    # 核心计算
    # ======================
    df["商品成本"] = -df["商品采购成本"] / df["销量"]
    df["商品售价"] = (df["商品优惠后金额（仅普通单）"] + df["其他优惠"]) / df["销量"]
    df["毛利预估"] = (df["商品售价"] * 0.8 - 5 - df["商品成本"]) / df["商品售价"]
    df["目标价"] = ((5 + df["商品成本"]) / 0.6).apply(math.ceil)
    df["建议调价幅度"] = (df["目标价"] - df["商品售价"]) / df["商品售价"]

    # ======================
    # 加载控价SKU数据
    # ======================
    price_control_sku_dict = {}
    if price_control_sku_file is not None:
        df_price_control = pd.read_excel(price_control_sku_file)
        if len(df_price_control.columns) < 2:
            st.error("❌ 控价SKU文件必须包含【系统SKU】和【控价价格】两列")
            st.stop()
        pc_sku_col = next((c for c in df_price_control.columns if c in ["系统SKU", "SKU", "商品SKU", "SKU编码"]), df_price_control.columns[0])
        pc_price_col = next((c for c in df_price_control.columns if "控价" in c or "价格" in c), df_price_control.columns[1])
        
        df_price_control[pc_sku_col] = df_price_control[pc_sku_col].astype(str).str.strip()
        df_price_control[pc_price_col] = pd.to_numeric(df_price_control[pc_price_col], errors='coerce')
        df_price_control = df_price_control.dropna(subset=[pc_price_col])
        price_control_sku_dict = df_price_control.set_index(pc_sku_col)[pc_price_col].to_dict()
        st.warning(f"💰 已加载控价SKU数量：{len(price_control_sku_dict)} 个")

    # ======================
    # 步骤1：按毛利阈值筛选
    # ======================
    df_result = df[df["毛利预估"] < profit_threshold].copy()

    # ======================
    # 步骤2：排除店铺
    # ======================
    exclude_shops = []
    if exclude_shop_file is not None:
        df_shop = pd.read_excel(exclude_shop_file)
        exclude_shops = df_shop.iloc[:, 0].astype(str).str.strip().tolist()
        st.warning(f"🚫 已加载排除店铺数量：{len(exclude_shops)} 个")
        df_result = df_result[df_result[shop_col].astype(str).str.strip().isin(exclude_shops) == False].copy()

    # ======================
    # 步骤3：排除SKU
    # ======================
    exclude_skus = []
    if exclude_sku_file is not None:
        df_sku = pd.read_excel(exclude_sku_file)
        exclude_skus = df_sku.iloc[:, 0].astype(str).str.strip().tolist()
        st.warning(f"🚫 已加载排除SKU数量：{len(exclude_skus)} 个")
        df_result = df_result[df_result[sku_col].astype(str).str.strip().isin(exclude_skus) == False].copy()

    # ======================
    # 步骤3.5：匹配控价 + 备注
    # ======================
    if price_control_sku_dict:
        df_result["系统SKU_匹配"] = df_result[sku_col].astype(str).str.strip()
        df_result["控价价格"] = df_result["系统SKU_匹配"].map(price_control_sku_dict)
        mask = df_result["控价价格"].notna()
        df_result.loc[mask, "目标价"] = df_result.loc[mask, "控价价格"].apply(math.ceil)
        df_result.loc[mask, "建议调价幅度"] = (df_result.loc[mask, "目标价"] - df_result.loc[mask, "商品售价"]) / df_result.loc[mask, "商品售价"]
        
        df_result["备注"] = ""
        df_result.loc[mask, "备注"] = "控价SKU"
        
        df_result = df_result.drop(columns=["系统SKU_匹配"])
        matched_count = mask.sum()
        st.success(f"🎯 匹配到控价价格的SKU数量：{matched_count} 个")
    else:
        df_result["备注"] = ""

    # ======================
    # 步骤4：列整理 + 排序
    # ======================
    output_cols = [shop_col, sku_col, "销量", "商品成本", "商品售价", "毛利预估", 
                   "控价价格", "目标价", "建议调价幅度", "备注"]
    df_result = df_result[output_cols].rename(columns={shop_col: "店铺", sku_col: "系统SKU"})
    df_result = df_result.sort_values(by="店铺", ascending=True).reset_index(drop=True)

    # ======================
    # 界面展示
    # ======================
    st.subheader("📈 最终筛选结果（按店铺排序）")
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("最终需提商品数", len(df_result))
    col2.metric("涉及店铺数", df_result["店铺"].nunique())
    col3.metric("已排除店铺数", len(exclude_shops))
    col4.metric("平均毛利", f"{df_result['毛利预估'].mean()*100:.2f}%")

    df_show = df_result.copy()
    df_show["毛利预估"] = df_show["毛利预估"].apply(lambda x: f"{x*100:.2f}%")
    df_show["建议调价幅度"] = df_show["建议调价幅度"].apply(lambda x: f"{x*100:.2f}%")
    df_show["商品成本"] = df_show["商品成本"].round(2)
    df_show["商品售价"] = df_show["商品售价"].round(2)
    df_show["控价价格"] = pd.to_numeric(df_show["控价价格"], errors='coerce').round(2)
    df_show["控价价格"] = df_show["控价价格"].fillna("-")
    df_show["目标价"] = df_show["目标价"].round(0).astype(int)

    st.dataframe(df_show, use_container_width=True, hide_index=True)

    # ======================
    # 下载Excel
    # ======================
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df_result.to_excel(writer, sheet_name="提价清单", index=False)
        wb = writer.book
        ws = wb.active

        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = center_align
                if col == 6:
                    cell.number_format = "0.00%"
                elif col == 9:
                    cell.number_format = "0.00%"
                elif col in [4,5,7]:
                    cell.number_format = "0.00"
                elif col in [3,8]:
                    cell.number_format = "0"

        for col in ws.columns:
            max_len = max(len(str(cell.value)) for cell in col if cell.value is not None)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 22)

    out.seek(0)

    st.download_button(
        "📥 下载最终提价报表（按店铺排序）",
        data=out,
        file_name="低毛利商品提价清单.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    st.info("👈 请先上传主Excel文件，即可开启所有筛选功能")
